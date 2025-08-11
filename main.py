from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime


def report_creating(input_file, column_name, value, output_file):
    try:
        workbook = load_workbook(rf"{input_file}")
        sheet = workbook.active

        # Determine the position of table in the sheet
        for row in sheet["A1:Z100"]:
            for cell in row:
                if cell.value == "№ п/п":
                    start_coord = cell.coordinate

        # Collecting coordinates of all cells in table header
        for row in sheet[f"{start_coord}:H9"]:
            params = {cell.value: cell.coordinate for cell in row}

        # Taking coordinates of required column
        column_coord = [
            value for key, value in params.items() if key.lower() == column_name.lower()
        ]
        if len(column_coord) == 0:
            raise ValueError(
                f"Column <{column_name}> is unknown for table <{input_file}>"
            )

        # Store of cells with appropriate value
        cells_positions = list()

        start = column_coord[0][0] + str(int(column_coord[0][1]) + 2)
        end = column_coord[0][0] + str(1000)

        for column in sheet[f"{start}:{end}"]:
            for cell in column:
                if cell.value is None:
                    break
                elif isinstance(value, int):
                    if cell.value == value:
                        cells_positions.append(cell.coordinate)
                elif isinstance(value, str):
                    if cell.value.lower() == value.lower():
                        cells_positions.append(cell.coordinate)

        if len(cells_positions) == 0:
            raise ValueError(
                f"Value <{value}> is inappropriate for column <{column_name}>"
            )

        result = [["ФИО", "Должность", "Отдел", "Дата найма", "Зарплата"]]

        for item in range(len(cells_positions)):
            array = list()
            result.append(array)
            for column in result[0]:
                for key, value in params.items():
                    if key == column:
                        # Retrieving cell values of Excel-File from input
                        cell_value = sheet[value[0] + cells_positions[item][1:]].value
                        array.append(cell_value)

        # Creating new Excel-File for output
        new_workbook = Workbook()
        sheet = new_workbook.active

        sheet["A1"] = "Тип:"
        sheet["B1"] = "Excel файл"
        sheet["A2"] = "Дата выгрузки:"
        sheet["B2"] = datetime.now().strftime("%d/%m/%Y")

        sheet["C5"] = "Результирующий набор сотрудников:"

        sheet["A1"].font = Font(bold=True)
        sheet["A2"].font = Font(bold=True)
        sheet["C5"].font = Font(bold=True)

        sheet["A8"] = None

        for row in result:
            sheet.append(row)

        table = Table(displayName="Report", ref=f"A9:E{len(cells_positions)+9}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
            showColumnStripes=True,
        )
        table.tableStyleInfo = style

        sheet.add_table(table)

        for letter in ["A", "B", "D", "E"]:
            sheet.column_dimensions[letter].width = 15
        sheet.column_dimensions["C"].width = 35

        new_workbook.save(rf"{output_file}")

    except Exception as e:
        return f"Error: {e}"
    else:
        return "success"
